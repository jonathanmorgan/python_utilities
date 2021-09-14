# TODO
# - add timing information - keep track of:
#     - // start
#     - // time and total elapsed at each periodic output.
#     - elapsed for just period
#     - average per item.

#===============================================================================
# imports
#===============================================================================


# base python libraries
import datetime
import dateutil
import logging
import openpyxl
import sys
import time
import traceback

# python_utilities
from python_utilities.status.status_container import StatusContainer

# ETL imports
from python_utilities.etl.etl_attribute import ETLAttribute
from python_utilities.etl.etl_entity import ETLEntity
from python_utilities.etl.etl_error import ETLError
from python_utilities.etl.etl_processor import ETLProcessor
from python_utilities.etl.etl_object_loader import ETLObjectLoader
#from python_utilities.etl.etl_django_model_loader import ETLDjangoModelLoader
from python_utilities.etl.etl_from_dictionary import ETLFromDictionary


#===============================================================================
# class ETLFromExcelWithHeaders
#===============================================================================


# lineage: object --> ETLProcessor --> ETLObjectLoader --> ETLDjangoModelLoader --> ETLFromDictionary
class ETLFromExcelWithHeaders( ETLFromDictionary ):


    #===========================================================================
    # CONSTANTS-ish
    #===========================================================================


    #STATUS_SUCCESS = "Success!"
    #STATUS_PREFIX_ERROR = "ERROR: "

    # logger name
    MY_LOGGER_NAME = "python_utilities.etl.ETLFromExcelWithHeaders"


    #===========================================================================
    # ! ==> class variables
    #===========================================================================


    # debug_flag
    debug_flag = False

    # logging
    #logging_level = logging.ERROR
    #update_status_every = 1000


    #===========================================================================
    # ! ==> class methods
    #===========================================================================


    #===========================================================================
    # ! ==> __init__() method - instance variables
    #===========================================================================


    def __init__( self ):

        '''
        Constructor
        '''

        # call parent's __init__()
        super().__init__()

        # spec
        #self.etl_entity = None

        # input
        self.input_worksheet = None

        # status - worksheet-level
        #self.status_message_list = []
        #self.latest_status = None
        #self.unknown_attrs_list = []
        #self.existing_count = 0
        #self.new_count = 0
        #self.start_dt = None

        # status - row-level
        #self.unknown_attrs_name_to_value_map = {}

        # debug
        self.debug_flag = False

    #-- END constructor --#


    #===========================================================================
    # ! ==> instance methods
    #===========================================================================


    def get_input_worksheet( self ):

        # return reference
        value_OUT = None

        # get value
        value_OUT = self.input_worksheet

        return value_OUT

    #-- END method get_input_worksheet() --#


    def get_value_for_key( self, record_IN, key_IN ):

        # return reference
        value_OUT = False

        # input parameters
        row_index_IN = None

        # declare variables
        me = "get_value_for_key"
        status_message = None
        my_debug_flag = None
        my_worksheet = None
        current_column_index = None
        current_cell = None
        current_value = None

        # first, check type of record.

        # ==> dictionary?
        if ( isinstance( record_IN, dict ) ):

            # dictionary - call parent method.
            value_OUT = super().get_value_for_key(
                record_IN,
                key_IN
            )

        elif ( isinstance( record_IN, int ) ):

            # init
            my_debug_flag = self.debug_flag
            row_index_IN = record_IN  # for this type, pass index, not record.

            # get spec information
            my_worksheet = self.get_input_worksheet()
            my_etl_spec = self.get_etl_entity()

            if ( my_debug_flag == True ):
                status_message = "Getting value for key: {}".format( key_IN )
                self.output_debug( status_message, method_IN = me, do_print_IN = my_debug_flag )
            #-- END DEBUG --#

            # retrieve column index
            current_column_index = my_etl_spec.pull_index_for_key( key_IN )

            if ( my_debug_flag == True ):
                status_message = "- Current column index: {}".format( current_column_index )
                self.output_debug( status_message, method_IN = me, do_print_IN = my_debug_flag )
            #-- END DEBUG --#

            if ( ( current_column_index is not None ) and ( current_column_index != "" ) ):

                # retrieve value from this row.
                current_cell = my_worksheet.cell( row = row_index_IN, column = current_column_index )
                current_value = current_cell.value
                value_OUT = current_value

                if ( my_debug_flag == True ):
                    status_message = "- Current column value: {}".format( current_value )
                    self.output_debug( status_message, method_IN = me, do_print_IN = my_debug_flag )
                #-- END DEBUG --#

            else:

                # ERROR
                status_message = "In method ETLFromExcelWithHeaders.{method_name}(): No index found for key {key_name}.".format(
                    method_name = me,
                    key_name = key_IN
                )
                self.output_log_message( status_message, me, log_level_code_IN = logging.ERROR, do_print_IN = True )

                # unknown key - can't get value.
                value_OUT = None

                raise ETLError( status_message )

            #-- END check to see if key maps to index. --#

        else:

            # ERROR
            status_message = "In method ETLFromExcelWithHeaders.{method_name}(): record is of unexpected type: {record_type}, not int or dict ( record: {record} ).".format(
                method_name = me,
                record_type = type( record_IN ),
                record = record_IN
            )
            self.output_log_message( status_message, me, log_level_code_IN = logging.ERROR, do_print_IN = True )

            # unknown record type - can't get value.
            value_OUT = None

            raise ETLError( status_message )

        #-- END check type of record passed in. --#

        return value_OUT

    #-- END method get_value_for_key() --#


    def map_indexes_to_keys( self ):

        # declare variables
        me = "map_indexes_to_keys"
        status_message = None
        my_debug_flag = None
        my_etl_spec = None
        my_worksheet = None
        my_class = None
        index_to_key_map = None
        key_to_index_map = None
        missing_attr_list = None

        # declare variables - column processing
        column_count = None
        header_row_number = None
        current_column_index = None
        current_cell = None
        current_column_name = None
        attr_exists = None

        #----------------------------------------------------------------------#
        # work

        my_debug_flag = self.debug_flag

        # get instances
        my_etl_spec = self.get_etl_entity()
        my_worksheet = self.get_input_worksheet()
        my_class = my_etl_spec.get_load_class()

        # create dictionaries in which we store maps of index to value and value to index for each column in first row.
        index_to_key_map = {}
        key_to_index_map = {}
        missing_attr_list = []

        # initialize processing
        column_count = my_worksheet.max_column
        header_row_number = 1

        # loop over columns to get value from first row for each.
        for current_column_index in range( 1, column_count + 1 ):

            # retrieve value for this cell
            current_cell = my_worksheet.cell( row = header_row_number, column = current_column_index )
            current_column_name = current_cell.value

            # store in index-to-name map
            index_to_key_map[ current_column_index ] = current_column_name
            key_to_index_map[ current_column_name ] = current_column_index

            # is this name in the instance?
            attr_exists = hasattr( my_class, current_column_name )

            if ( my_debug_flag == True ):
                status_message = "index {index}; name = {name} ( exists?: {exists_flag} )".format(
                    index = current_column_index,
                    name = current_column_name,
                    exists_flag = attr_exists
                )
                self.output_debug( status_message, method_IN = me, do_print_IN = my_debug_flag )
            #-- END DEBUG --#

            # if name not in instance, add to list for processing.
            if ( attr_exists == False ):

                # attr does not exist.
                missing_attr_list.append( current_column_name )

            #-- END check if has attr --#

        #-- END loop over columns in 1st row --#

        # store the maps.
        my_etl_spec.set_attr_index_to_key_map( index_to_key_map )
        my_etl_spec.set_attr_key_to_index_map( key_to_index_map )

        # store unknown attributes.
        self.set_unknown_attrs_list( missing_attr_list )

        status_message = "fields that don't correspond to attrs in instance: {}".format( missing_attr_list )
        self.add_status_message( status_message )

    #-- END method map_indexes_to_keys() --#


    def set_input_worksheet( self, value_IN ):

        # return reference
        value_OUT = None

        # declare variables
        my_worksheet = None
        row_count = None
        first_data_row_index = None
        end_data_row_index = None
        row_index_list = None
        row_index_iterator = None

        # init
        first_data_row_index = 2

        # store value
        self.input_worksheet = value_IN

        # clear out status variables.
        self.reset_status_information()

        # return value
        value_OUT = self.get_input_worksheet()

        # get row count and create iterator over row indexes within this
        #     worksheet, starting with row 2 (skipping header row).
        my_worksheet = value_OUT

        # get row count...
        row_count = my_worksheet.max_row

        # ...make list of row indexes...
        first_data_row_index = 2
        end_data_row_index = row_count
        row_index_list = range( first_data_row_index, end_data_row_index + 1 )

        # ...and store list (from which iterator can be made each time it is
        #     needed).
        self.set_record_list( row_index_list )

        return value_OUT

    #-- END method set_input_worksheet() --#


    def update_instance_from_record( self, instance_IN, record_IN, save_on_success_IN = True ):

        # return reference
        status_OUT = None

        # declare variables
        me = "update_instance_from_record"
        my_debug_flag = None
        status_message = None
        record_dictionary = None
        my_etl_spec = None
        current_entry_instance = None
        current_row_index = None
        my_worksheet = None
        row_count = None
        column_count = None
        current_column_index = None
        current_cell = None
        current_column_value = None
        current_column_key = None
        current_attr_name = None
        current_attr_value = None
        my_etl_attribute = None

        #----------------------------------------------------------------------#
        # ==> do work

        # init
        status_OUT = StatusContainer()
        status_OUT.set_status_code( StatusContainer.STATUS_CODE_SUCCESS )
        my_debug_flag = self.debug_flag
        my_etl_spec = self.get_etl_entity()
        current_entry_instance = instance_IN
        current_row_index = record_IN
        record_dictionary = {}

        # store supporting information in status instance.
        status_OUT = self.init_status( status_OUT )

        # got an instance?
        if ( current_entry_instance is not None ):

            # get worksheet
            my_worksheet = self.get_input_worksheet()
            row_count = my_worksheet.max_row
            column_count = my_worksheet.max_column

            #------------------------------------------------------------------#
            # ==> process record attributes (loop over columns in row)

            # loop over columns to get values for each column and create a
            #     dictionary that we'll pass on to parent method.
            for current_column_index in range( 1, column_count + 1 ):

                #--------------------------------------------------------------#
                # ==> get attribute spec, name and value

                # value --> retrieve value for this cell
                current_cell = my_worksheet.cell( row = current_row_index, column = current_column_index )
                current_column_value = current_cell.value

                # key --> get key for index.
                current_column_key = my_etl_spec.pull_key_for_index( current_column_index )

                # start with attribute value set to column value
                current_attr_value = current_column_value

                # retrieve ETLAttribute for this index.
                my_etl_attribute = my_etl_spec.pull_attr_for_key( current_column_key )

                # got an attribute spec?
                if ( my_etl_attribute is not None ):

                    # ETLAttribute found.  Retrieve values and do
                    #     processing.
                    current_attr_name = my_etl_attribute.get_load_attr_name()

                else:

                    # no ETLAttribute instance. Direct, untranslated
                    #     load to instance.
                    current_attr_name = current_column_key

                #-- END check to see if attribute spec. --#

                # add attribute_name and value to record dictionary.
                record_dictionary[ current_attr_name ] = current_attr_value

            #-- END loop over cells in spreadsheet row. --#

            #--------------------------------------------------------------#
            # ==> get attribute spec, name and value

            # call parent method to process spreadsheet row as dictionary, so
            #     we get all of the pre- and post-processing there built-in.
            status_OUT = super().update_instance_from_record(
                current_entry_instance,
                record_dictionary
            )

        else:

            # no instance passed in - log error, return error status.
            status_message = "ERROR - No instance passed in to update. Can't do anything."
            self.output_log_message( status_message, method_IN = me, log_level_code_IN = logging.ERROR )
            self.add_status_message( status_message )

            # status
            status_OUT.set_status_code( StatusContainer.STATUS_CODE_ERROR )
            status_OUT.add_message( status_message )
            status_OUT.set_detail_value( self.PROP_WAS_INSTANCE_UPDATED, False )

        #-- END check to see if instance is not None --#

        return status_OUT

    #-- END method update_instance_from_record() --#


#-- END class ETLFromExcelWithHeaders --#
